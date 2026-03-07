from rest_framework.decorators import api_view, permission_classes, throttle_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle

from django.contrib.auth import authenticate, login, logout, get_user_model
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.utils.text import slugify
from django.db.models import Q
from django.db import transaction

from datetime import datetime, date
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from django.http import HttpResponse

from rest_framework.authtoken.models import Token
from collections import defaultdict
import httpx
import requests
import uuid
from urllib.parse import quote
import difflib
import re

from .models import (
    User,
    Department,
    Issue,
    IssueDepartment,
    Minutes,
    Response as ResponseModel,
    Notification
)
from .services import check_and_send_overdue_emails
from .serializers import IssueDepartmentSerializer, NotificationSerializer, DPOIssueSerializer
from .gemini_utils import analyze_document_with_gemini, match_issues_with_gemini
from .supabase_utils import upload_to_supabase
import os
from datetime import datetime, date, timedelta
import io
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ================= SUPABASE STORAGE CONFIG ================= #
SUPABASE_URL = os.getenv('SUPABASE_URL', '')
SUPABASE_KEY = os.getenv('SUPABASE_KEY', '')
SUPABASE_BUCKET = os.getenv('SUPABASE_BUCKET_NAME', 'Mnutes')

# ================= TEMP CACHE (ONLY FOR NOW) ================= #
TEMP_DATA_CACHE = []


class LoginRateThrottle(AnonRateThrottle):
    scope = 'login'


# ================= AUTH ================= #


@api_view(['POST'])
@permission_classes([AllowAny]) 
@throttle_classes([LoginRateThrottle])
def login_view(request):
    username_input = (request.data.get('username') or '').strip()
    password_input = request.data.get('password') or ''

    if not username_input or not password_input:
        return Response({"success": False, "message": "Missing credentials"}, status=400)
    user = authenticate(username=username_input, password=password_input)

    if user is None or not user.is_active:
        return Response({"success": False, "message": "Invalid username or password"}, status=401)

    Token.objects.filter(user=user).delete()
    token = Token.objects.create(user=user)

    return Response({
        "success": True,
        "token": token.key,
        "username": user.username,
        "role": (user.role or '').lower(),
        "department": user.department.dept_name if user.department else None
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    Token.objects.filter(user=request.user).delete()
    return Response({"success": True})


def _resolve_department_username(dept_name):
    base = slugify(dept_name)[:40].replace('-', '_') or 'department'
    candidate = base
    suffix = 1

    while User.objects.filter(username=candidate).exists():
        suffix += 1
        suffix_text = f'_{suffix}'
        candidate = f'{base[:150 - len(suffix_text)]}{suffix_text}'

    return candidate


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_department_user(request):
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Only DPO can add department users"}, status=403)

    dept_name = (request.data.get('dept_name') or '').strip()
    designation = (request.data.get('designation') or '').strip()
    email = (request.data.get('email') or '').strip()
    password = request.data.get('password') or ''
    confirm_password = request.data.get('confirm_password') or ''

    if not dept_name:
        return Response({"error": "Department name is required"}, status=400)
    if not designation:
        return Response({"error": "Designation is required"}, status=400)
    if not email:
        return Response({"error": "Department email is required"}, status=400)
    if not password:
        return Response({"error": "Password is required"}, status=400)
    if password != confirm_password:
        return Response({"error": "Passwords do not match"}, status=400)
    if len(password) < 6:
        return Response({"error": "Password must be at least 6 characters"}, status=400)

    if Department.objects.filter(dept_name__iexact=dept_name).exists():
        return Response({"error": "Department already exists"}, status=409)

    username = _resolve_department_username(dept_name)

    with transaction.atomic():
        department = Department.objects.create(
            dept_name=dept_name,
            designation=designation,
            email=email,
        )

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            role='department',
            department=department,
            is_active=True,
        )

    return Response({
        "success": True,
        "department": department.dept_name,
        "designation": department.designation,
        "username": user.username,
    }, status=201)



# ================= MINUTES UPLOAD ================= #

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_minutes(request):
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Only DPO can upload minutes"}, status=403)

    global TEMP_DATA_CACHE

    file = request.FILES.get('file')
    meeting_date_str = request.POST.get('meeting_date')
    
    if not file:
        return Response({"error": "File required"}, status=400)

    # Save file locally temporarily for Gemini analysis
    os.makedirs('media/minutes', exist_ok=True)
    local_file_path = f"media/minutes/{file.name}"

    with open(local_file_path, 'wb+') as dest:
        for chunk in file.chunks():
            dest.write(chunk)

    # Build a list of formatted department strings: "Designation, Department Name"
    available_departments = []
    for dept in Department.objects.all():
        desig = (dept.designation or '').strip()
        name = (dept.dept_name or '').strip()
        if desig:
            available_departments.append(f"{desig}, {name}")
        else:
            available_departments.append(name)

    # Analyze with Gemini using current department master list (including designations)
    raw_data = analyze_document_with_gemini(local_file_path, available_departments)

    # Upload to Supabase Storage using shared utility
    with open(local_file_path, 'rb') as f:
        file_data = f.read()
    supabase_public_url = upload_to_supabase(
        file_data=file_data,
        original_filename=file.name,
        bucket=SUPABASE_BUCKET,
        folder='public'
    )

    dept_objects = list(Department.objects.all())
    dept_name_map = _build_department_name_maps(dept_objects)

    print(f"🔍 Gemini returned {len(raw_data)} issues.")
    
    clean_data = []
    for i, item in enumerate(raw_data):
        print(f"--- Processing Issue {i+1} ---")
        item = _normalize_issue_departments(item, dept_name_map)
        print(f"  Matched Depts: {item.get('departments')}")
        clean_data.append(item)

    # ===== CREATE MINUTES RECORD NOW (not waiting for allocate_all) ===== #
    u_by = request.user
    # Parse meeting date if provided
    try:
        if meeting_date_str:
            parts = meeting_date_str.split('-')
            if len(parts) == 3:
                meeting_date_obj = datetime.strptime(meeting_date_str, '%d-%m-%Y').date()
            else:
                meeting_date_obj = timezone.now().date()
        else:
            meeting_date_obj = timezone.now().date()
    except:
        meeting_date_obj = timezone.now().date()

    # Use Supabase URL if upload succeeded, otherwise fall back to local path
    file_path_for_db = supabase_public_url or local_file_path
    
    # Create Minutes record IMMEDIATELY
    minute_obj = Minutes.objects.create(
        title=file.name,  # Use actual filename as title
        meeting_date=meeting_date_obj,
        uploaded_by=u_by,
        file_path=file_path_for_db
    )
    
    print(f"✅ Minutes record created: ID {minute_obj.id}, Title: {file.name}")
    # ===== END NEW ===== #

    # Store file path and minutes ID in TEMP_DATA_CACHE for the allocate step
    TEMP_DATA_CACHE = {
        'file_path': supabase_public_url or local_file_path,
        'is_supabase': supabase_public_url is not None,
        'issues': clean_data,
        'minutes_id': minute_obj.id
    }
    
    return Response({"success": True, "data": clean_data, "minutes_id": minute_obj.id})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_assign_issues(request):
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Unauthorized"}, status=403)
    # Return only issues, not the file_path
    if isinstance(TEMP_DATA_CACHE, dict) and 'issues' in TEMP_DATA_CACHE:
        return Response(TEMP_DATA_CACHE['issues'])
    return Response(TEMP_DATA_CACHE)


# ================= ISSUE ALLOCATION ================= #

def _resolve_minutes_id(request):
    request_minutes_id = request.data.get('minutes_id') or request.data.get('minutesId')
    try:
        minutes_id = int(request_minutes_id) if request_minutes_id not in [None, '', 'null', 'undefined'] else None
    except (TypeError, ValueError):
        minutes_id = None

    if minutes_id is None and isinstance(TEMP_DATA_CACHE, dict):
        minutes_id = TEMP_DATA_CACHE.get('minutes_id')

    return minutes_id


def _resolve_minutes_obj(minutes_id):
    if not minutes_id:
        return None, Response({"error": "minutes_id is required for allocation."}, status=400)

    try:
        minute_obj = Minutes.objects.get(id=minutes_id)
        return minute_obj, None
    except Minutes.DoesNotExist:
        return None, Response({"error": "Minutes record not found. Please upload file again."}, status=400)


def _parse_deadline(item):
    """Parse deadline from item data, defaulting to 14 days from today."""
    raw = item.get('deadline', '')
    if raw and raw.strip():
        try:
            # Handle DD-MM-YYYY format from frontend
            parts = raw.strip().split('-')
            if len(parts) == 3:
                return date(int(parts[2]), int(parts[1]), int(parts[0]))
        except (ValueError, IndexError):
            pass
    return date.today() + timedelta(days=14)


def _extract_departments(item):
    val = item.get('departments')
    # AI now returns a list of objects: [{"designation": "...", "department": "..."}, ...]
    if isinstance(val, list):
        return val
    
    # Fallback for old/manual data
    val = item.get('department', 'GENERAL')
    return [
        d.strip()[:100]
        for d in str(val).split(',')
        if d.strip()
    ]


def _dept_key(value):
    # Keep alphabetical and numeric characters from any language (Malayalam, etc)
    return re.sub(r'[^\w]+', '', (value or '').strip().upper())


def _build_department_name_maps(departments):
    exact = {}
    normalized = {}
    designation_map = {}
    combined = {}

    for dept in departments:
        name = (dept.dept_name or '').strip()
        desig = (dept.designation or '').strip()
        if not name:
            continue

        exact[name.upper()] = dept
        
        n_name = _dept_key(name)
        if n_name and n_name not in normalized:
            normalized[n_name] = dept
            
        if desig:
            n_desig = _dept_key(desig)
            if n_desig and n_desig not in designation_map:
                designation_map[n_desig] = dept
            
            n_comb = _dept_key(f"{desig}{name}")
            if n_comb and n_comb not in combined:
                combined[n_comb] = dept

    return {
        'exact': exact,
        'normalized': normalized,
        'designation': designation_map,
        'combined': combined,
        'normalized_keys': list(normalized.keys()),
    }


def _match_department(raw_input, dept_maps, cutoff=0.72):
    if not raw_input:
        return None

    # Handle structured object from Gemini: {"designation": "...", "department": "..."}
    if isinstance(raw_input, dict):
        d_val = (raw_input.get('designation') or '').strip()
        n_val = (raw_input.get('department') or '').strip()
        query = f"{d_val}{n_val}"
        raw_name = f"{d_val}, {n_val}" if d_val and n_val else (d_val or n_val)
    else:
        query = raw_input
        raw_name = raw_input

    print(f"    Matching Dept: '{raw_name}'")
    
    exact = dept_maps['exact']
    normalized = dept_maps['normalized']
    combined = dept_maps.get('combined', {})
    designation_map = dept_maps.get('designation', {})

    # 1. Try exact name match if query is a string
    if not isinstance(raw_input, dict):
        exact_match = exact.get(query.strip().upper())
        if exact_match:
            return exact_match

    # 2. Try normalized combined match (designation + name)
    key = _dept_key(query)
    if not key:
        return None

    if key in combined:
        return combined[key]

    # 3. Try normalized name match
    if key in normalized:
        return normalized[key]
        
    # 4. Try normalized designation match (only if raw_input was a string/designation only)
    if key in designation_map:
        return designation_map[key]

    # 5. Handle alias-like values by containment
    containment_candidates = []
    for normalized_key, dept in normalized.items():
        if key in normalized_key or normalized_key in key:
            containment_candidates.append((abs(len(normalized_key) - len(key)), dept))
            
    for comb_key, dept in combined.items():
        if key in comb_key:
             containment_candidates.append((abs(len(comb_key) - len(key)), dept))

    if containment_candidates:
        containment_candidates.sort(key=lambda x: x[0])
        return containment_candidates[0][1]

    # 6. Fuzzy match fallback
    close = difflib.get_close_matches(key, dept_maps['normalized_keys'], n=1, cutoff=0.55)
    if close:
        return normalized.get(close[0])

    print(f"    ⚠️ Match Failed for: '{raw_name}'")
    return None


def _normalize_issue_departments(item, dept_maps):
    raw_depts = item.get('departments') or []
    if not raw_depts:
        raw_depts = [item.get('department', '')]

    resolved = []
    unresolved = []
    seen = set()

    for raw_input in raw_depts:
        if not raw_input:
            continue

        matched = _match_department(raw_input, dept_maps)
        if matched:
            canonical = matched.dept_name
            if canonical not in seen:
                resolved.append(canonical)
                seen.add(canonical)
        else:
            # For unresolved, if it's a dict, convert to a readable string for display
            if isinstance(raw_input, dict):
                d_val = (raw_input.get('designation') or '').strip()
                n_val = (raw_input.get('department') or '').strip()
                label = f"{d_val}, {n_val}" if d_val and n_val else (d_val or n_val)
                unresolved.append(label or "Unknown")
            else:
                unresolved.append(str(raw_input))

    if resolved:
        item['departments'] = resolved
        item['department'] = ', '.join(resolved)
    else:
        item['departments'] = unresolved
        item['department'] = ', '.join(unresolved)

    return item


def _collect_unknown_departments(issues_data):
    dept_maps = _build_department_name_maps(Department.objects.all())
    unknown = set()

    for item in issues_data:
        for dept_input in _extract_departments(item):
            if not _match_department(dept_input, dept_maps):
                if isinstance(dept_input, dict):
                    d_val = (dept_input.get('designation') or '').strip()
                    n_val = (dept_input.get('department') or '').strip()
                    label = f"{d_val}, {n_val}" if d_val and n_val else (d_val or n_val)
                    unknown.add(label or "Unknown Stakeholder")
                else:
                    unknown.add(str(dept_input))

    return sorted(list(unknown))


def _create_issues_for_minutes(minute_obj, issues_data):
    dept_counts = {}
    dept_maps = _build_department_name_maps(Department.objects.all())

    for item in issues_data:
        issue_title = item.get('issue', 'No Title')

        # Resolve parent_issue if provided
        parent_issue_id = item.get('parent_issue_id')
        parent_issue_obj = None
        if parent_issue_id:
            try:
                parent_issue_obj = Issue.objects.get(id=int(parent_issue_id))
                # Star topology: always point to the absolute root
                # Recursive traversal to the top-most parent
                visited = set()
                while parent_issue_obj.parent_issue and parent_issue_obj.id not in visited:
                    visited.add(parent_issue_obj.id)
                    parent_issue_obj = parent_issue_obj.parent_issue
            except (Issue.DoesNotExist, ValueError, TypeError):
                parent_issue_obj = None

        # Use update_or_create to avoid duplicates when re-allocating from drafts.
        # Match on minutes + issue_title to find existing issues.
        issue, created = Issue.objects.update_or_create(
            minutes=minute_obj,
            issue_title=issue_title,
            defaults={
                'issue_no': str(item.get('issue_no', '')),
                'issue_description': item.get('issue_description', ''),
                'location': item.get('location', ''),
                'priority': item.get('priority', 'Medium'),
                'parent_issue': parent_issue_obj,
            }
        )

        normalized_item = _normalize_issue_departments(item, dept_maps)
        dept_list = _extract_departments(normalized_item)

        d_date = _parse_deadline(item)

        for dept_name in dept_list:
            dept = _match_department(dept_name, dept_maps)
            if not dept:
                continue

            # Use update_or_create for IssueDepartment too, so re-allocation
            # updates the existing assignment instead of creating a duplicate.
            _issue_dept, id_created = IssueDepartment.objects.update_or_create(
                issue=issue,
                department=dept,
                defaults={
                    'deadline_date': d_date,
                    'status': 'pending',
                }
            )

            # Only count as new assignment if the IssueDepartment was just created
            if id_created:
                dept_counts[dept] = dept_counts.get(dept, 0) + 1

    # Only send notifications for genuinely new assignments
    for dept, count in dept_counts.items():
        users = User.objects.filter(department=dept).exclude(role__in=['DPO', 'COLLECTOR'])
        for u in users:
            Notification.objects.create(
                user=u,
                issue_department=None,
                message=f"ACTION REQUIRED: {count} new issues have been assigned to your department."
            )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def allocate_single(request):
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Only DPO can allocate"}, status=403)

    issue_item = request.data.get('issue')
    if not isinstance(issue_item, dict):
        return Response({"error": "issue payload is required."}, status=400)

    minutes_id = _resolve_minutes_id(request)
    minute_obj, err = _resolve_minutes_obj(minutes_id)
    if err:
        return err

    unknown_departments = _collect_unknown_departments([issue_item])
    if unknown_departments:
        return Response(
            {
                "error": "Unknown department(s). Please map to master department list.",
                "unknown_departments": unknown_departments,
            },
            status=400,
        )

    _create_issues_for_minutes(minute_obj, [issue_item])
    return Response({"success": True, "allocated": 1})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def allocate_all(request):
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Only DPO can allocate"}, status=403)

    global TEMP_DATA_CACHE

    issues_data = request.data.get('issues', [])
    if not issues_data and isinstance(TEMP_DATA_CACHE, dict):
        issues_data = TEMP_DATA_CACHE.get('issues', [])
    elif not issues_data and isinstance(TEMP_DATA_CACHE, list):
        issues_data = TEMP_DATA_CACHE

    if not isinstance(issues_data, list) or len(issues_data) == 0:
        return Response({"error": "No issues provided for allocation."}, status=400)

    minutes_id = _resolve_minutes_id(request)
    minute_obj, err = _resolve_minutes_obj(minutes_id)
    if err:
        return err

    unknown_departments = _collect_unknown_departments(issues_data)
    if unknown_departments:
        return Response(
            {
                "error": "Unknown department(s). Please map to master department list.",
                "unknown_departments": unknown_departments,
            },
            status=400,
        )

    _create_issues_for_minutes(minute_obj, issues_data)

    TEMP_DATA_CACHE = []
    return Response({"success": True, "allocated": len(issues_data)})


# ================= ISSUES ================= #

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_all_issues(request):
    if request.user.role.lower() not in ['dpo', 'collector']:
        return Response({"error": "Unauthorized"}, status=403)

    today = date.today()
    
    # 1. Update Overdue Status (Optimized)
    overdue_assignments = IssueDepartment.objects.filter(
        status__iexact='pending', 
        deadline_date__lt=today
    )
    for i in overdue_assignments:
        i.status = 'overdue'
        i.save()

    # 2. Get the date filter
    filter_date = request.query_params.get('date')

    # 3. Start Query (FIXED for Speed AND Accuracy)
    # We kept 'department' (Major Speed Boost) but REMOVED 'responses'.
    # This ensures the serializer finds the latest response text correctly.
    issues_query = Issue.objects.select_related(
        'minutes'
    ).prefetch_related(
        'issuedepartment_set',
        'issuedepartment_set__department'
    ).all().order_by('-id')

    # 4. Apply Date Filter
    if filter_date:
        issues_query = issues_query.filter(minutes__meeting_date__date=filter_date)

    serializer = DPOIssueSerializer(issues_query, many=True)
    return Response(serializer.data)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_dept_issues(request, dept_name):
    if request.user.role.lower() != 'department':
        return Response({"error": "Unauthorized"}, status=403)

    if not request.user.department:
        return Response({"error": "Department mapping missing for user"}, status=403)

    requested_dept_name = (dept_name or '').strip().lower()
    user_dept_name = (request.user.department.dept_name or '').strip().lower()
    if requested_dept_name != user_dept_name:
        return Response({"error": "Unauthorized department access"}, status=403)

    today = date.today()
    
    # 1. OPTIMIZATION: 'select_related' fetches the parent Issue and Dept info in the SAME query.
    # 'prefetch_related' fetches all the responses in the SAME query.
    issues_query = IssueDepartment.objects.select_related('issue', 'department').prefetch_related('responses').filter(
        department=request.user.department
    ).order_by('-issue__id') 
    
    # 2. Update Overdue Status (Only for pending items to save time)
    # We iterate through the already fetched list to avoid a second DB hit
    for i in issues_query:
        if i.status.lower() == 'pending' and i.deadline_date and i.deadline_date < today:
            i.status = 'overdue'
            i.save()
            
    serializer = IssueDepartmentSerializer(issues_query, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def submit_response(request):
    if request.user.role.lower() != 'department':
        return Response({"error": "Only department users can submit responses"}, status=403)

    if not request.user.department:
        return Response({"error": "Department mapping missing for user"}, status=403)

    # 1. Get data safely (handle both JSON and FormData)
    issue_id = request.data.get('issue_id') or request.data.get('id')
    response_text = request.data.get('response')
    attachment = request.FILES.get('attachment')

    print(f"📝 Submitting response for ID: {issue_id}")

    if not issue_id or str(issue_id) == "undefined":
        return Response({"error": "Invalid ID provided"}, status=400)

    try:
        issue_link = IssueDepartment.objects.get(
            id=issue_id,
            department=request.user.department,
        )
        
        # Upload attachment to Supabase 'Response' bucket
        attachment_url = None
        if attachment:
            file_bytes = attachment.read()
            attachment_url = upload_to_supabase(
                file_data=file_bytes,
                original_filename=attachment.name,
                bucket='Response',
                folder='public'
            )
            if not attachment_url:
                print(f"⚠️ Supabase upload failed for attachment, falling back to local storage")
                # Rewind for local save fallback
                attachment.seek(0)

        # Save Response — use URL if Supabase succeeded, otherwise local file
        ResponseModel.objects.create(
            issue_department=issue_link,
            response_text=response_text or "",
            attachment_path=attachment_url if attachment_url else (attachment or None)
        )

        issue_link.status = 'submitted'
        issue_link.save()
        
        # --- NOTIFY DPO ---
        dpos = User.objects.filter(Q(role__iexact='DPO') | Q(username__iexact='dpo'))
        issue_number = issue_link.issue.id 
        dept_name = issue_link.department.dept_name

        for d in dpos:
            Notification.objects.create(
                user=d, 
                issue_department=issue_link,
                message=f"Response Received: {dept_name} responded to Issue #{issue_number}"
            )
            
        return Response({"success": True})

    except IssueDepartment.DoesNotExist:
        return Response({"error": "Issue not found or not assigned to your department"}, status=404)
    except Exception as e:
        print(f"🔥 CRITICAL SUBMIT ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_notifications(request):
    notifs = Notification.objects.filter(
        user=request.user
    ).order_by('-created_at')[:20]

    return Response(NotificationSerializer(notifs, many=True).data)

# ================= REPORTS ================= #


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_report(request):
    # 1. Parse request data sent from the frontend modal
    req_format = request.data.get('format', 'excel')
    reports_data = request.data.get('reports', [])

    # Fallback just in case no data was sent (backward compatibility)
    if not reports_data:
        return Response({"error": "No report data provided."}, status=400)

    # ==========================================
    #             PDF GENERATION
    # ==========================================
    if req_format == 'pdf':
        buffer = io.BytesIO()
        # Set to Landscape A4 for TV Presentation
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        elements = []
        styles = getSampleStyleSheet()
        
        # --- REGISTER MALAYALAM FONT ---
        import os
        from django.conf import settings
        
        # Adjust path if you placed the font somewhere else
        font_path = os.path.join(settings.BASE_DIR, 'font', 'NotoSansMalayalam-Regular.ttf') 
        
        try:
            pdfmetrics.registerFont(TTFont('MalayalamFont', font_path))
            font_name = 'MalayalamFont'
        except Exception as e:
            print(f"Font loading error: {e}")
            font_name = 'Helvetica' # Fallback if font isn't found

        # Presentation Styles
        issue_no_style = ParagraphStyle(
            'IssueNo', parent=styles['Normal'], fontName=font_name, fontSize=18,
            textColor=colors.white, backColor=colors.HexColor('#1E3A8A'),
            borderPadding=8, spaceAfter=20
        )
        issue_text_style = ParagraphStyle(
            'IssueText', parent=styles['Normal'], fontName=font_name, fontSize=24,
            leading=34, spaceAfter=40
        )
        dept_style = ParagraphStyle(
            'DeptText', parent=styles['Normal'], fontName=font_name, fontSize=18,
            textColor=colors.red, alignment=2 # 2 = Right aligned
        )
        response_header_style = ParagraphStyle(
            'ResponseHeader', parent=styles['Normal'], fontName=font_name, fontSize=22,
            textColor=colors.white, backColor=colors.HexColor('#0284c7'), # Lighter blue
            borderPadding=10, spaceAfter=20
        )
        response_text_style = ParagraphStyle(
            'ResponseText', parent=styles['Normal'], fontName=font_name, fontSize=22,
            leading=32, spaceAfter=20
        )

        for report in reports_data:
            # --- SLIDE 1: THE ISSUE ---
            issue_title = report.get('issue_no', 'Unknown Issue')
            elements.append(Paragraph(f"Issue: {issue_title}", issue_no_style))
            
            issue_desc = report.get('issue', '')
            elements.append(Paragraph(issue_desc, issue_text_style))
            
            dept_name = report.get('department', '')
            elements.append(Paragraph(f"{dept_name}", dept_style))
            
            # --- SLIDE 2: THE RESPONSE ---
            elements.append(PageBreak()) 
            
            elements.append(Paragraph("നിലവിലെ അവസ്ഥ / Action Taken", response_header_style))
            
            response_text = report.get('response', 'No response yet')
            # Split response by newlines so it formats nicely
            for para in response_text.split('\n'):
                if para.strip():
                    elements.append(Paragraph(para.strip(), response_text_style))
                    elements.append(Spacer(1, 10))

            # Move to next issue
            elements.append(PageBreak())

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="DDC_Presentation.pdf"'
        response.write(pdf)
        return response

    # ==========================================
    #            EXCEL GENERATION
    # ==========================================
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Follow Up Report"

        header_font = Font(bold=True, size=11, name='Calibri')
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        content_align = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.column_dimensions['A'].width = 10  
        ws.column_dimensions['B'].width = 40  
        ws.column_dimensions['C'].width = 30  
        ws.column_dimensions['D'].width = 80  
        ws.column_dimensions['E'].width = 60  

        headers = [
            "ക്രമ നമ്പർ", 
            "ഉന്നയിച്ച തീയതി & വകുപ്പ്/വിഷയം", 
            "നടപടി സ്വീകരിക്കേണ്ട ഉദ്യോഗസ്ഥൻ",
            "മുൻ യോഗത്തിൽ ചർച്ച ചെയ്തതും യോഗ നിർദ്ദേശവും", 
            "നിലവിലെ സ്റ്റാറ്റസ്"
        ]
        
        # Default if no dates are found
        meeting_date_str = timezone.now().strftime("%d-%m-%Y")
        
        # Try to get the meeting date from the first issue for the header
        first_report = reports_data[0] if reports_data else {}
        raw_meeting_date = first_report.get('meeting_date')
        if raw_meeting_date:
            try:
                # Expecting YYYY-MM-DD from serializer
                if 'T' in raw_meeting_date:
                    raw_meeting_date = raw_meeting_date.split('T')[0]
                dt = datetime.strptime(raw_meeting_date, '%Y-%m-%d')
                meeting_date_str = dt.strftime("%d-%m-%Y")
            except:
                pass
        ws.merge_cells('A1:E1')
        main_header = ws['A1']
        main_header.value = f"{meeting_date_str} -ന് ചേർന്ന ജില്ലാ വികസന സമിതി യോഗത്തിന്റെ തുടർ നടപടി റിപ്പോർട്ട്"
        main_header.font = Font(bold=True, size=12, name='Calibri')
        main_header.alignment = Alignment(horizontal='center', vertical='center')
        main_header.border = thin_border
        ws.row_dimensions[1].height = 30

        ws.append(headers)

        for cell in ws[2]:
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            ws.row_dimensions[2].height = 45

        # Populate Excel with the EDITED data from the frontend
        for index, report in enumerate(reports_data, start=1):
            # Use specific meeting date for this issue
            row_meeting_date_str = meeting_date_str # Fallback to header date
            raw_row_date = report.get('meeting_date')
            if raw_row_date:
                try:
                    if 'T' in raw_row_date: raw_row_date = raw_row_date.split('T')[0]
                    dt = datetime.strptime(raw_row_date, '%Y-%m-%d')
                    row_meeting_date_str = dt.strftime("%d-%m-%Y")
                except:
                    pass

            subject_col_text = f"തീയതി: {row_meeting_date_str}\n\n{report.get('issue_no', '')}"

            row_data = [
                index,
                subject_col_text,
                report.get('department', ''),
                report.get('issue', ''),
                report.get('response', '')
            ]
            ws.append(row_data)

            current_row = ws.max_row
            for cell in ws[current_row]:
                cell.alignment = content_align
                cell.border = thin_border

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Follow_Up_Report.xlsx"'
        wb.save(response)
        return response

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def send_overdue_alerts(request):
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Unauthorized"}, status=403)
    
    sent_count = check_and_send_overdue_emails()
    
    return Response({
        "success": True,
        "sent_count": sent_count
    })


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def resolve_issue(request, issue_id):
    """Toggle resolution_status between resolved/unresolved."""
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Only DPO can resolve issues"}, status=403)

    try:
        issue = Issue.objects.get(id=issue_id)
    except Issue.DoesNotExist:
        return Response({"error": "Issue not found"}, status=404)

    new_status = request.data.get('resolution_status')
    if new_status not in ('resolved', 'unresolved'):
        return Response({"error": "Invalid status. Use 'resolved' or 'unresolved'."}, status=400)

    issue.resolution_status = new_status
    issue.save(update_fields=['resolution_status'])

    return Response({
        "success": True,
        "id": issue.id,
        "resolution_status": issue.resolution_status
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_issue_lifecycle(request, issue_id):
    """Return lifecycle chain for a given issue with optional filters."""
    if request.user.role.lower() not in ['dpo', 'collector']:
        return Response({"error": "Unauthorized"}, status=403)

    try:
        target_issue = Issue.objects.select_related('parent_issue').get(id=issue_id)
    except Issue.DoesNotExist:
        return Response({"error": "Issue not found"}, status=404)

    # Find the absolute root
    root_issue = target_issue
    visited = set()
    while root_issue.parent_issue and root_issue.id not in visited:
        visited.add(root_issue.id)
        root_issue = root_issue.parent_issue

    # Collect all issues in the chain (Root + all descendants)
    # Since chains are small, we can do this in a few steps or use a broader filter
    all_related_ids = {root_issue.id}
    
    # Simple iterative approach to find descendants (since chains are likely < 10 items)
    current_parents = [root_issue.id]
    for _ in range(10): # Max depth safeguard
        descendants = list(Issue.objects.filter(parent_issue_id__in=current_parents).values_list('id', flat=True))
        if not descendants:
            break
        all_related_ids.update(descendants)
        current_parents = descendants

    lifecycle_qs = Issue.objects.select_related(
        'minutes', 'parent_issue'
    ).prefetch_related(
        'issuedepartment_set',
        'issuedepartment_set__department',
        'issuedepartment_set__responses'
    ).filter(id__in=all_related_ids)

    serialized_items = list(DPOIssueSerializer(lifecycle_qs, many=True).data)

    status_filter = (request.query_params.get('status') or '').strip().lower()
    department_filter = (request.query_params.get('department') or '').strip().lower()
    search_filter = (request.query_params.get('search') or '').strip().lower()
    from_date = (request.query_params.get('from_date') or '').strip()
    to_date = (request.query_params.get('to_date') or '').strip()
    has_response = (request.query_params.get('has_response') or '').strip().lower()

    def normalized_status(value):
        s = str(value or 'pending').lower().strip()
        if s in ['submitted', 'completed', 'received']:
            return 'received'
        if s == 'overdue':
            return 'overdue'
        return 'pending'

    def in_date_range(item_date):
        if not item_date:
            return not (from_date or to_date)
        date_part = str(item_date)[:10]
        try:
            item_dt = datetime.strptime(date_part, '%Y-%m-%d').date()
        except ValueError:
            return True

        if from_date:
            try:
                from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
                if item_dt < from_dt:
                    return False
            except ValueError:
                pass

        if to_date:
            try:
                to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
                if item_dt > to_dt:
                    return False
            except ValueError:
                pass

        return True

    filtered_items = []
    for item in serialized_items:
        item_status = normalized_status(item.get('status'))
        item_department = str(item.get('department') or '').lower()
        item_search_blob = ' '.join([
            str(item.get('issue') or ''),
            str(item.get('issue_description') or ''),
            str(item.get('minutes_title') or ''),
            str(item.get('issue_no') or ''),
        ]).lower()
        responses = item.get('response') or []

        if status_filter and status_filter != 'all' and item_status != status_filter:
            continue
        if department_filter and department_filter != 'all' and department_filter not in item_department:
            continue
        if search_filter and search_filter not in item_search_blob:
            continue
        if not in_date_range(item.get('meeting_date')):
            continue
        if has_response == 'true' and len(responses) == 0:
            continue
        if has_response == 'false' and len(responses) > 0:
            continue

        filtered_items.append(item)

    filtered_items.sort(key=lambda item: (
        str(item.get('meeting_date') or ''),
        str(item.get('created_at') or ''),
        int(item.get('id') or 0),
    ))

    return Response({
        'issue_id': target_issue.id,
        'issue_no': target_issue.issue_no,
        'root_issue_id': root_issue.id,
        'root_issue_no': root_issue.issue_no,
        'total_iterations': len(serialized_items),
        'filtered_iterations': len(filtered_items),
        'items': filtered_items,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_existing_issues(request):
    """Return all existing issues for the flag panel on the allocation page."""
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Unauthorized"}, status=403)

    issues = Issue.objects.select_related('minutes').filter(resolution_status='unresolved').order_by('-id')

    result = []
    for iss in issues:
        # Get departments for this issue
        depts = ", ".join(
            a.department.dept_name
            for a in iss.issuedepartment_set.select_related('department').all()
        )
        result.append({
            'id': iss.id,
            'issue': iss.issue_title,
            'issue_description': iss.issue_description,
            'issue_no': iss.issue_no,
            'minutes_title': iss.minutes.title if iss.minutes else '',
            'minutes_id': iss.minutes.id if iss.minutes else None,
            'meeting_date': iss.minutes.meeting_date.isoformat() if iss.minutes and iss.minutes.meeting_date else None,
            'department': depts,
            'location': iss.location,
        })

    return Response(result)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def match_issues(request):
    """Use Gemini to suggest matches between new and existing issues."""
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Unauthorized"}, status=403)

    new_issues = request.data.get('new_issues', [])
    existing_issues = request.data.get('existing_issues', [])

    if not new_issues or not existing_issues:
        return Response([])

    matches = match_issues_with_gemini(new_issues, existing_issues)
    return Response(matches)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_minutes(request):
    """Fetch all uploaded minutes - accessible to DPO and Department users"""
    # Allow both DPO and department users to view
    if request.user.role.lower() not in ['dpo', 'department', 'collector']:
        return Response({"error": "Unauthorized to view minutes"}, status=403)
    
    minutes = Minutes.objects.all().order_by('-created_at').prefetch_related('issues')
    
    minutes_data = []
    for m in minutes:
        # Convert FieldFile to string
        file_path_str = str(m.file_path) if m.file_path else None
        
        # file_path is already a full URL if from Supabase, or a local path if not
        file_url = file_path_str
        if file_url and not file_url.startswith('http'):
            # Local file - add /media/ prefix
            file_url = f"/media/{file_url}"
        
        # Extract original filename from the file path
        # For Supabase files: "public/uuid_filename.pdf" -> "filename.pdf"
        # For local files: "media/minutes/filename.pdf" -> "filename.pdf"
        original_filename = file_path_str
        if original_filename and '_' in original_filename:
            # Extract filename after UUID prefix
            original_filename = original_filename.split('_', 1)[1]
        elif original_filename:
            # Extract just the filename from path
            original_filename = original_filename.split('/')[-1]
        
        minutes_data.append({
            'id': m.id,
            'title': m.title,
            'originalFileName': original_filename,
            'uploadedDate': m.created_at.isoformat(),
            'meetingDate': m.meeting_date.isoformat() if m.meeting_date else None,
            'fileUrl': file_url,
            'uploadedBy': m.uploaded_by.username if m.uploaded_by else 'Unknown',
            'issueCount': m.issues.count()
        })
    
    return Response(minutes_data)


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_minutes(request, minutes_id):
    """Delete a Minutes record"""
    if request.user.role.lower() != 'dpo':
        return Response({"error": "Only DPO can delete minutes"}, status=403)
    
    try:
        minute = Minutes.objects.get(id=minutes_id)
        
        # Get original filename for response
        file_path_str = str(minute.file_path) if minute.file_path else "Unknown"
        original_filename = file_path_str
        if original_filename and '_' in original_filename:
            original_filename = original_filename.split('_', 1)[1]
        elif original_filename:
            original_filename = original_filename.split('/')[-1]
        
        print(f"🗑️  Deleting Minutes: {original_filename} (ID: {minutes_id})")
        
        # Delete the minute and associated issues
        minute.delete()
        
        return Response({
            "success": True,
            "message": f"Deleted: {original_filename}"
        })
    
    except Minutes.DoesNotExist:
        return Response({"error": "Minutes record not found"}, status=404)
    except Exception as e:
        print(f"❌ Error deleting minutes: {e}")
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_minutes_pdf(request):
    try:
        from weasyprint import HTML
    except ImportError as e:
        print("WeasyPrint error:", e)
        return Response({"error": "PDF renderer (WeasyPrint) not installed."}, status=500)

    html_content = request.data.get('html', '')
    if not html_content:
        return Response({"error": "No HTML content provided"}, status=400)
    
    # Quill CSS mapped to WeasyPrint
    quill_css = """
    @page { 
        size: A4 portrait; 
        margin: 1.5cm 2.0cm; 
    }
    body { 
        font-family: 'Manjari', 'Meera', 'Noto Sans Malayalam', 'Times New Roman', Times, serif; 
        font-size: 14pt; 
        line-height: 1.6;
    }
    .minute-title-section { text-align: center; font-weight: bold; margin-bottom: 30px; }
    .ql-align-center { text-align: center; }
    .ql-align-right { text-align: right; }
    .ql-align-justify { text-align: justify; }
    .ql-indent-1 { padding-left: 3em; }
    .ql-indent-2 { padding-left: 6em; }
    .ql-indent-3 { padding-left: 9em; }
    .ql-indent-4 { padding-left: 12em; }
    .ql-indent-5 { padding-left: 15em; }
    .ql-indent-6 { padding-left: 18em; }
    .ql-indent-7 { padding-left: 21em; }
    .ql-indent-8 { padding-left: 24em; }
    ul, ol { padding-left: 1.5em; margin-top: 5px; margin-bottom: 5px; }
    li { margin-bottom: 5px; }
    p { margin-top: 0; margin-bottom: 5px; }
    strong { font-weight: bold; }
    em { font-style: italic; }
    u { text-decoration: underline; }
    """

    full_html = f"""
    <!DOCTYPE html>
    <html lang="ml">
    <head>
        <meta charset="utf-8">
        <style>{quill_css}</style>
    </head>
    <body>
        <div class="ql-editor">
            {html_content}
        </div>
    </body>
    </html>
    """

    try:
        pdf_file = HTML(string=full_html).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="minutes.pdf"'
        return response
    except Exception as e:
        print("WeasyPrint conversion error:", e)
        return Response({"error": str(e)}, status=500)
